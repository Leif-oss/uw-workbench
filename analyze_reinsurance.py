import openpyxl
import sys
import io

# Set stdout to handle UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('Reinsurance calculator - FInal (1).xlsx')

print("Sheet names:", wb.sheetnames)
print("\n" + "="*80)

for sheet_name in wb.sheetnames:
    print(f"\n\n=== SHEET: {sheet_name} ===\n")
    ws = wb[sheet_name]
    
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
    print("\nFirst 30 rows:\n")
    
    for i, row in enumerate(ws.iter_rows(values_only=False), 1):
        row_values = []
        for cell in row:
            if cell.value is not None:
                # Show formula if it exists
                if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                    row_values.append(f"{cell.coordinate}: FORMULA={cell.value}")
                else:
                    row_values.append(f"{cell.coordinate}: {cell.value}")
        
        if row_values:  # Only print non-empty rows
            print(f"Row {i}: {' | '.join(row_values)}")
        
        if i >= 30:
            break

wb.close()

